import asyncio
import uuid
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from app.core.database import SessionLocal, get_db
from app.core.deps import get_admin_user
from app.models.book import Book
from app.models.essay import Essay, EssayTopic
from app.models.mcq import Mcq, McqSet
from app.models.user import User
from app.schemas.common import Msg
from app.schemas.essay import EssayTopicCreate
from app.services.file_service import NOTES_DIR

router = APIRouter(prefix="/admin", tags=["admin"])

# ── Scraper ────────────────────────────────────────────────────────────────────

_scrape_jobs: dict[str, dict[str, Any]] = {}

# Each subject maps to one or more category slugs on pakmcqs.com
SUBJECT_URLS: dict[str, list[str]] = {
    "Current Affairs":                ["pakistan-current-affairs-mcqs", "world-current-affairs-mcqs"],
    "Pakistan Affairs":               ["pak-study-mcqs"],
    "Islamic Studies":                ["islamic-studies-mcqs"],
    "General Science & Ability":      ["everyday-science-mcqs"],
    "English (Precis & Composition)": ["english-mcqs"],
    "General Knowledge":              ["general_knowledge_mcqs"],
}

_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}


def _extract_mcq(mcq_soup: Any) -> tuple[str, str, str, str, str, str]:
    """Extract question, four options, and correct letter from a pakmcqs.com detail page."""
    question_tag = mcq_soup.find("h1")
    if not question_tag:
        return "", "", "", "", "", ""
    question_text = question_tag.get_text(strip=True)
    options: dict[str, str] = {"A": "", "B": "", "C": "", "D": ""}
    correct_answer = ""
    p_tag = question_tag.find_next("p")
    if p_tag:
        for child in p_tag.children:
            if isinstance(child, str):
                for line in child.split("\n"):
                    for letter in "ABCD":
                        if line.startswith(f"{letter}."):
                            options[letter] = line[2:].strip()
            elif getattr(child, "name", None) == "strong":
                strong_text = child.get_text(strip=True)
                for letter in "ABCD":
                    if strong_text.startswith(f"{letter}."):
                        options[letter] = strong_text[2:].strip()
                        correct_answer = letter
    return question_text, options["A"], options["B"], options["C"], options["D"], correct_answer


def _scrape_sync(job_id: str, subject: str, pages: int) -> None:
    """Run in a thread executor — scrapes pakmcqs.com and saves directly to DB."""
    import requests
    from bs4 import BeautifulSoup

    slugs = SUBJECT_URLS[subject]  # one or more category slugs
    # pages are split evenly across slugs; e.g. 10 pages / 2 slugs = 5 pages each
    pages_per_slug = max(1, pages // len(slugs))
    total_pages = pages_per_slug * len(slugs)
    _scrape_jobs[job_id]["total_pages"] = total_pages

    db = SessionLocal()
    try:
        mcq_set = db.query(McqSet).filter(McqSet.subject == subject).first()
        if not mcq_set:
            mcq_set = McqSet(subject=subject, source_file="scraped")
            db.add(mcq_set)
            db.flush()

        existing: set[str] = {
            row[0] for row in db.query(Mcq.question).filter(Mcq.set_id == mcq_set.id).all()
        }

        added = 0
        errors = 0
        global_page = 0

        for slug in slugs:
            base_url = f"https://pakmcqs.com/category/{slug}/page/"
            for page_num in range(1, pages_per_slug + 1):
                global_page += 1
                _scrape_jobs[job_id]["page"] = global_page
                try:
                    resp = requests.get(f"{base_url}{page_num}", timeout=15, headers=_HEADERS)
                    resp.raise_for_status()
                except Exception as e:
                    errors += 1
                    _scrape_jobs[job_id]["errors"] = errors
                    _scrape_jobs[job_id]["last_error"] = str(e)
                    continue

                soup = BeautifulSoup(resp.text, "html.parser")
                links = soup.find_all("a", class_="read-more-link read-more-basic")

                for link in links:
                    mcq_url = link.get("href", "")
                    if not mcq_url:
                        continue
                    try:
                        mcq_resp = requests.get(mcq_url, timeout=15, headers=_HEADERS)
                        mcq_resp.raise_for_status()
                        mcq_soup = BeautifulSoup(mcq_resp.text, "html.parser")
                        q, a, b, c, d, correct = _extract_mcq(mcq_soup)
                        if not q or q in existing or not any([a, b, c, d]):
                            continue
                        db.add(Mcq(
                            set_id=mcq_set.id,
                            question=q, option_a=a, option_b=b, option_c=c, option_d=d,
                            correct=correct or "A",
                        ))
                        existing.add(q)
                        added += 1
                        _scrape_jobs[job_id]["added"] = added
                        if added % 20 == 0:
                            db.commit()
                    except Exception as e:
                        errors += 1
                        _scrape_jobs[job_id]["errors"] = errors
                        _scrape_jobs[job_id]["last_error"] = str(e)

        db.commit()
        _scrape_jobs[job_id].update({"status": "done", "added": added, "errors": errors})

    except Exception as exc:
        db.rollback()
        _scrape_jobs[job_id].update({"status": "error", "message": str(exc)})
    finally:
        db.close()


@router.post("/mcqs/scrape")
async def start_scrape(
    subject: str = Form(...),
    pages: int = Form(5),
    _: User = Depends(get_admin_user),
):
    if subject not in SUBJECT_URLS:
        raise HTTPException(400, f"Unknown subject. Valid: {list(SUBJECT_URLS.keys())}")
    job_id = str(uuid.uuid4())
    _scrape_jobs[job_id] = {
        "status": "running", "added": 0, "errors": 0,
        "page": 0, "total_pages": pages, "subject": subject,
    }
    loop = asyncio.get_event_loop()
    loop.run_in_executor(None, _scrape_sync, job_id, subject, pages)
    return {"job_id": job_id, "detail": f"Scraping started for {subject}"}


@router.get("/mcqs/scrape/status/{job_id}")
async def scrape_status(job_id: str, _: User = Depends(get_admin_user)):
    job = _scrape_jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    return job


@router.get("/stats")
def stats(db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    return {
        "users": db.query(User).count(),
        "essays": db.query(Essay).count(),
        "essays_pending": db.query(Essay).filter(Essay.status == "pending").count(),
        "mcqs": db.query(Mcq).count(),
        "books": db.query(Book).count(),
    }


@router.get("/users")
def list_users(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=5, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    q = db.query(User)
    total = q.count()
    users = q.order_by(User.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": [
            {
                "id": u.id,
                "name": u.name,
                "email": u.email,
                "is_admin": u.is_admin,
                "is_active": u.is_active,
                "created_at": u.created_at.isoformat(),
                "prep_level": u.profile.prep_level if u.profile else None,
                "exam_type": u.profile.exam_type if u.profile else None,
            }
            for u in users
        ],
        "total": total,
    }


@router.patch("/users/{user_id}/toggle", response_model=Msg)
def toggle_user(user_id: int, db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, "User not found")
    user.is_active = not user.is_active
    db.commit()
    return Msg(detail=f"User {'activated' if user.is_active else 'deactivated'}")


@router.delete("/users/{user_id}", response_model=Msg)
def delete_user(user_id: int, db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, "User not found")
    db.delete(user)
    db.commit()
    return Msg(detail="User deleted")


@router.post("/mcqs/upload", response_model=Msg)
async def upload_mcqs(
    subject: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    import openpyxl, io
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]

    existing = db.query(McqSet).filter(McqSet.subject == subject).first()
    if existing:
        db.delete(existing)
        db.flush()

    mcq_set = McqSet(subject=subject, source_file=file.filename)
    db.add(mcq_set)
    db.flush()

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("question"):
            continue
        rows.append(
            Mcq(
                set_id=mcq_set.id,
                question=str(data.get("question", "")),
                option_a=str(data.get("option_a", data.get("a", ""))),
                option_b=str(data.get("option_b", data.get("b", ""))),
                option_c=str(data.get("option_c", data.get("c", ""))),
                option_d=str(data.get("option_d", data.get("d", ""))),
                correct=str(data.get("correct", "A")).strip().upper()[0],
            )
        )

    db.bulk_save_objects(rows)
    db.commit()
    return Msg(detail=f"Imported {len(rows)} MCQs for {subject}")


@router.post("/notes/upload", response_model=Msg)
async def upload_note(
    file: UploadFile = File(...),
    _: User = Depends(get_admin_user),
):
    NOTES_DIR.mkdir(parents=True, exist_ok=True)
    content = await file.read()
    dest = NOTES_DIR / (file.filename or "note.pdf")
    dest.write_bytes(content)
    return Msg(detail=f"Uploaded {file.filename}")


@router.delete("/notes/{filename}", response_model=Msg)
def delete_note(filename: str, _: User = Depends(get_admin_user)):
    path = NOTES_DIR / filename
    if not path.exists():
        raise HTTPException(404, "File not found")
    path.unlink()
    return Msg(detail="Deleted")


@router.post("/topics", status_code=201)
def create_topic(data: EssayTopicCreate, db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    topic = EssayTopic(category=data.category, title=data.title)
    db.add(topic)
    db.commit()
    return {"id": topic.id, "detail": "Created"}


@router.patch("/topics/{topic_id}", response_model=Msg)
def update_topic(topic_id: int, data: EssayTopicCreate, db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    topic = db.get(EssayTopic, topic_id)
    if not topic:
        raise HTTPException(404, "Topic not found")
    topic.category = data.category
    topic.title = data.title
    db.commit()
    return Msg(detail="Updated")


@router.delete("/topics/{topic_id}", response_model=Msg)
def delete_topic(topic_id: int, db: Session = Depends(get_db), _: User = Depends(get_admin_user)):
    topic = db.get(EssayTopic, topic_id)
    if not topic:
        raise HTTPException(404, "Topic not found")
    db.delete(topic)
    db.commit()
    return Msg(detail="Deleted")
