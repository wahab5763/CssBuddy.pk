"""
Import MCQs from xlsx files in the project /data folder.
Usage: cd backend && .venv\Scripts\python scripts/seed_mcqs_from_data.py
"""
import sys, os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
os.chdir(Path(__file__).parent.parent)

import openpyxl
from app.core.database import SessionLocal
from app.models.mcq import Mcq, McqSet

DATA_DIR = Path(__file__).parent.parent.parent / "data"

SUBJECT_MAP = {
    "Current Affairs.xlsx":  "Current Affairs",
    "English.xlsx":          "English (Precis & Composition)",
    "GSA.xlsx":              "General Science & Ability",
    "Islamic Studies.xlsx":  "Islamic Studies",
    "Pak Affairs.xlsx":      "Pakistan Affairs",
}


def seed():
    db = SessionLocal()
    try:
        total_imported = 0
        for filename, subject in SUBJECT_MAP.items():
            path = DATA_DIR / filename
            if not path.exists():
                print(f"  SKIP: {filename} not found at {path}")
                continue

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Normalise headers: lowercase + underscores
            raw_headers = [str(c.value or "").strip() for c in ws[1]]
            headers = [h.lower().replace(" ", "_") for h in raw_headers]
            print(f"  {filename}: headers = {raw_headers}")

            # Remove existing set for this subject
            existing = db.query(McqSet).filter(McqSet.subject == subject).first()
            if existing:
                db.delete(existing)
                db.flush()

            mcq_set = McqSet(subject=subject, source_file=filename)
            db.add(mcq_set)
            db.flush()

            rows, skipped = [], 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                q = data.get("question") or data.get("questions")
                if not q:
                    skipped += 1
                    continue
                correct_raw = str(data.get("correct_option") or data.get("correct") or "A").strip()
                # Correct Option column contains "A","B","C","D" — take first char
                correct = correct_raw[0].upper() if correct_raw else "A"

                rows.append(Mcq(
                    set_id=mcq_set.id,
                    question=str(q).strip(),
                    option_a=str(data.get("option_a") or "").strip(),
                    option_b=str(data.get("option_b") or "").strip(),
                    option_c=str(data.get("option_c") or "").strip(),
                    option_d=str(data.get("option_d") or "").strip(),
                    correct=correct,
                ))

            db.bulk_save_objects(rows)
            db.commit()
            total_imported += len(rows)
            print(f"  OK: {len(rows)} MCQs imported for '{subject}' ({skipped} skipped)")

        print(f"\nTotal: {total_imported} MCQs imported across {len(SUBJECT_MAP)} subjects.")
    finally:
        db.close()


if __name__ == "__main__":
    print("Seeding MCQs from /data folder...")
    seed()
    print("Done.")
