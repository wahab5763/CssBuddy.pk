"""
Seed MCQs from Excel files in backend/data/excel/
Usage: cd backend && python scripts/seed_mcqs.py
"""
import sys
import os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
os.chdir(Path(__file__).parent.parent)

import openpyxl
from app.core.database import SessionLocal
from app.models.mcq import Mcq, McqSet

SUBJECT_MAP = {
    "current_affairs.xlsx": "Current Affairs",
    "english.xlsx": "English (Precis & Composition)",
    "gsa.xlsx": "General Science & Ability",
    "islamic_studies.xlsx": "Islamic Studies",
    "pakistan_affairs.xlsx": "Pakistan Affairs",
}

EXCEL_DIR = Path("data/excel")


def seed():
    db = SessionLocal()
    try:
        for filename, subject in SUBJECT_MAP.items():
            path = EXCEL_DIR / filename
            if not path.exists():
                print(f"  SKIP: {filename} not found")
                continue

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]

            existing = db.query(McqSet).filter(McqSet.subject == subject).first()
            if existing:
                db.delete(existing)
                db.flush()

            mcq_set = McqSet(subject=subject, source_file=filename)
            db.add(mcq_set)
            db.flush()

            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                q = data.get("question")
                if not q:
                    continue
                rows.append(
                    Mcq(
                        set_id=mcq_set.id,
                        question=str(q),
                        option_a=str(data.get("option_a", data.get("a", "")) or ""),
                        option_b=str(data.get("option_b", data.get("b", "")) or ""),
                        option_c=str(data.get("option_c", data.get("c", "")) or ""),
                        option_d=str(data.get("option_d", data.get("d", "")) or ""),
                        correct=str(data.get("correct", "A") or "A").strip().upper()[0],
                    )
                )

            db.bulk_save_objects(rows)
            db.commit()
            print(f"  OK: {len(rows)} MCQs imported for '{subject}'")

    finally:
        db.close()


if __name__ == "__main__":
    print("Seeding MCQs...")
    seed()
    print("Done.")
